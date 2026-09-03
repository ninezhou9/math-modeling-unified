#!/usr/bin/env python3
"""使用隔离的 LibreOffice 进程重算 XLSX 公式并检查公式错误。"""

import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# DOCX 与 XLSX 共用同一套 OOXML/LibreOffice 基础工具，避免两份代码漂移。
SHARED_OFFICE_SCRIPTS = Path(__file__).resolve().parents[2] / "docx" / "scripts"
sys.path.insert(0, str(SHARED_OFFICE_SCRIPTS))
from office.soffice import get_soffice_env


EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


def _error(message: str) -> dict:
    return {"status": "error", "error": message}


def _inspect_workbook(path: Path) -> dict:
    error_details = {error: [] for error in EXCEL_ERRORS}
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    for error in EXCEL_ERRORS:
                        if error in value:
                            error_details[error].append(f"{sheet.title}!{cell.coordinate}")
                            break
    finally:
        workbook.close()

    formula_count = 0
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        for sheet in workbook.worksheets:
            formula_count += sum(
                1
                for row in sheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            )
    finally:
        workbook.close()

    total_errors = sum(len(locations) for locations in error_details.values())
    return {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": formula_count,
        "error_summary": {
            error: {"count": len(locations), "locations": locations[:20]}
            for error, locations in error_details.items()
            if locations
        },
    }


def recalc(filename, timeout=30):
    """重算工作簿；任何超时、非零退出或缺失输出都不会覆盖原文件。"""
    source = Path(filename).resolve()
    if not source.exists():
        return _error(f"文件不存在: {source}")
    if source.suffix.lower() != ".xlsx":
        return _error(f"只支持 .xlsx 文件: {source}")

    try:
        with tempfile.TemporaryDirectory(prefix="math-modeling-lo-") as temp_dir:
            temp_root = Path(temp_dir)
            output_dir = temp_root / "output"
            output_dir.mkdir()
            profile_uri = (temp_root / "profile").resolve().as_uri()
            command = [
                "soffice",
                "--headless",
                "--norestore",
                f"-env:UserInstallation={profile_uri}",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(output_dir),
                str(source),
            ]
            completed = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=timeout,
                check=False,
                env=get_soffice_env(),
            )
            if completed.returncode != 0:
                detail = (completed.stderr or completed.stdout or "未知错误").strip()
                return _error(f"LibreOffice 重算失败: {detail}")

            converted = output_dir / source.name
            if not converted.exists():
                return _error("LibreOffice 未生成重算后的工作簿")

            result = _inspect_workbook(converted)
            replacement = source.with_name(f".{source.name}.recalc.tmp")
            shutil.copy2(converted, replacement)
            os.replace(replacement, source)
            return result
    except subprocess.TimeoutExpired:
        return _error(f"LibreOffice 重算超时（{timeout} 秒），原文件未修改")
    except FileNotFoundError:
        return _error("未找到 LibreOffice 可执行文件 soffice")
    except Exception as exc:
        return _error(f"重算失败，原文件未修改: {exc}")


def main() -> int:
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help", "/?"):
        print("用法: python recalc.py <工作簿.xlsx> [超时秒数]")
        print(
            "用隔离的 LibreOffice 进程重算 XLSX 公式并检查公式错误；"
            "任何失败或超时都不会覆盖原文件。"
        )
        return 0 if args else 2
    if args[0] == "--dry-run":
        if len(args) != 2:
            print("用法: python recalc.py --dry-run <工作簿.xlsx>", file=sys.stderr)
            return 2
        source = Path(args[1]).resolve()
        if not source.exists():
            result = _error(f"文件不存在: {source}")
        elif source.suffix.lower() != ".xlsx":
            result = _error(f"只支持 .xlsx 文件: {source}")
        else:
            try:
                result = _inspect_workbook(source)
                result["status"] = "dry_run"
            except Exception as error:
                result = _error(f"无效或损坏的 XLSX 文件: {error}")
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 1 if "error" in result else 0
    if len(args) > 2:
        print("用法: python recalc.py <工作簿.xlsx> [超时秒数]", file=sys.stderr)
        return 2
    if len(args) > 1:
        try:
            timeout = int(args[1])
        except ValueError:
            result = _error("timeout 必须为正整数")
            print(json.dumps(result, ensure_ascii=False, indent=2))
            return 1
        if timeout <= 0:
            result = _error("timeout 必须为正整数")
            print(json.dumps(result, ensure_ascii=False, indent=2))
            return 1
    else:
        timeout = 30
    result = recalc(args[0], timeout)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 1 if "error" in result else 0


if __name__ == "__main__":
    raise SystemExit(main())
